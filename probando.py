from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, BUILTIN_FORMATS
from constants import *
import time
start_time = time.time()


# 1. Crear new Workbook/Excel
wb = Workbook()
ws = wb.active
ws.title = "BASE PLAN MES N"
max_row = ws.max_row
max_col = ws.max_column

# 2. Agregamos los titulos de las columnas
ws.append(['Llave',
           'Sector',                        
           'Oficina',                       
           'Material',
           'Descripción',
           'Nivel Jer. 2',
           'Nivel Jer. 3',
           'RV Producción mes n+1',
           'RV Venta mes n+1',
           '% Uti. producción',
           'Producción disponible',
           'Stock al día',
           'Por producir mes N',
           'Producción por despachar mes N',
           'Total disponible',
           'Delay',
           'Vol. prom. Por contenedor',
           'Atraso a facturar',
           'Ajuste atraso',
           'Facturación atraso',
           'Producción para venta nueva',
           'Venta del mes',
           'Ajuste venta nueva',
           'Facturación Venta nueva',
           'Saldo Volumen disponible',
           'Disponible stock sin venta',
           'Ajuste stock sin venta',
           'Facturación stock',
           'En puerto a facturar',
           'Plan Irrestricto (DATO)',
           'Plan Ajustado',
           'Motivo Ajuste'
           ])

## 2.1 Para los estilos
for col in range(1, 33):
  ws[get_column_letter(col) + '1'].font = Font(bold=True, color=white)  # tomamos la primera fila
  nombre = ws.cell(row = 1, column = col).value
  ws.column_dimensions[get_column_letter(col)].width = tamano[nombre]
  ws.row_dimensions[1].height = 25

for col in range(1, 18):
   ws[get_column_letter(col) + '1'].fill = PatternFill("solid", fgColor=grey)

for col in range(18, 29, 4):
  ws[get_column_letter(col) + '1'].fill = PatternFill("solid", fgColor=lightBlue)
  ws[get_column_letter(col + 1) + '1'].fill = PatternFill("solid", fgColor=lightBlue)
  ws[get_column_letter(col + 2) + '1'].fill = PatternFill("solid", fgColor=orange)
  ws[get_column_letter(col + 3) + '1'].fill = PatternFill("solid", fgColor=grey)

ws[get_column_letter(29) + '1'].fill = PatternFill("solid", fgColor=orange)
ws[get_column_letter(30) + '1'].fill = PatternFill("solid", fgColor=blue)
ws[get_column_letter(31) + '1'].fill = PatternFill("solid", fgColor=blue)
ws[get_column_letter(32) + '1'].fill = PatternFill("solid", fgColor=blue)

for colu in range(8, 32):
  letter = str(get_column_letter(colu))
  if letter != 'J' or letter != 'AF':
    col = ws.column_dimensions[letter]
    col.number_format = BUILTIN_FORMATS[3]

col = ws.column_dimensions['K']
col.number_format = BUILTIN_FORMATS[3]
# 3. Agregamos la data

  ## ASIGNACIONES: Llave, Oficina, Material (=SKU), Descripción,  RV Producción mes n+1 (=kilos) 
asignacion = load_workbook(filename_asignacion)
  ## sheetnames: ['Producción', 'Venta']
produccion = asignacion['Producción']
max_rows_produccion = produccion.max_row
max_row_new = ws.max_row
max_rows = max_row_new

## Llave, Oficina, Sku, Descripción, RV producción mes n+1
for row in range(2, max_rows_produccion + 1):
  llave = produccion.cell(row = row, column = 1).value
  oficina = produccion.cell(row = row, column = 2).value
  sku = produccion.cell(row = row, column = 3).value
  descripcion = produccion.cell(row = row, column = 4).value
  prd_mes = produccion.cell(row = row, column = 5).value
  ws.cell(row = row, column = 1).value = llave
  ws.cell(row = row, column = 3).value = oficina
  ws.cell(row = row, column = 4).value = sku
  ws.cell(row = row, column = 5).value = descripcion
  ws.cell(row = row, column = 8).value = prd_mes

  # Colores rojo ajustes
  ws[f'S{row}'].font = Font(bold=True, color=red)
  ws[f'W{row}'].font = Font(bold=True, color=red)

  # Formato numeros
  ws[f'H{row}'].number_format = BUILTIN_FORMATS[3]
  ws[f'I{row}'].number_format = BUILTIN_FORMATS[3]

  if 'PV' in descripcion:
    ws.cell(row = row, column = 2).value = PV                                     # Sector = 'Pavo'       
  
  elif 'PO' in descripcion:
    ws.cell(row = row, column = 2).value = PO # Sector = 'Pollo'                               
  
  elif 'GO' in descripcion:
    ws.cell(row = row, column = 2).value = GO                                     # Sector = 'Cerdo'        

  elif 'GA' in descripcion:
    ws.cell(row = row, column = 2).value = GA # Sector = ''
  
  else:
    ws.cell(row = row, column = 2).value = 'Elaborado' # Sector = 'Elaborado'

asignacion.close()
wb.save(filename)


################## RV Venta mes n+1 ##################
plan_irrestricto = load_workbook(filename)
ws_irrestricto = plan_irrestricto.active
max_row = ws_irrestricto.max_row

venta = asignacion['Venta']
max_rows_venta = venta.max_row
dict_venta_n1 = {}

for row in range(2, max_rows_venta + 1):
  llave = venta.cell(row = row, column = 1).value
  venta_mes = venta.cell(row = row, column = 5).value
  dict_venta_n1[llave.lower()] = venta_mes

for search_row in range(2, max_row + 1):
    llave_actual = ws_irrestricto.cell(row = search_row, column = 1).value.lower()

    if llave_actual in dict_venta_n1:
      ws_irrestricto.cell(row = search_row, column = 9).value = dict_venta_n1[llave_actual]
  
    else:
      ws_irrestricto.cell(row = search_row, column = 9).value = 0 

plan_irrestricto.save(filename)

################## % Util. produccion ##################
################## Por producir mes N ##################
############ Producción por despachar mes N ############
######################## Delay #########################
################ Volumen por contenedor ################
################# En puerto a facturar #################
util_prod = load_workbook(filename_util, data_only=True, read_only=True)
ws_util = util_prod['Ponderación']
max_row_util = ws_util.max_row
util_dicc = {}

wb_stock = load_workbook(filename_stock, data_only=True, read_only=True)
ws_stock = wb_stock['TD Stock']
max_row_stock = ws_stock.max_row
dict_stock = {}

por_producir = load_workbook(filename_por_producir, data_only=True, read_only=True)
ws_por_producir = por_producir['Consolidado planificación']
ws_ponderaciones = por_producir['Ponderación cumplimiento']
max_row_por_producir = ws_por_producir.max_row
dict_producir = {}
dict_ponderacion_cumplimiento = {}

por_despachar = load_workbook(filename_por_despachar, data_only=True, read_only=True)
ws_por_despachar = por_despachar.active
max_row_por_despachar = ws_por_despachar.max_row
dict_por_despachar = {}

wb_delay = load_workbook(filename_delay, data_only=True, read_only=True)
ws_delay = wb_delay.active
max_row_delay = ws_delay.max_row
dict_delay = {}

wb_vol_cont = load_workbook(filename_vol_cont, data_only=True, read_only=True)
ws_contenedor = wb_vol_cont.active
max_row_cont = ws_contenedor.max_row
dict_vol_cont = {}


print("--- %s ANTES0 ---" % (time.time() - start_time))
wb_puerto = load_workbook(filename_puerto, data_only=True, read_only=True)
ws_puerto = wb_puerto.active
max_row_puerto = ws_puerto.max_row
dict_puerto = {}

print("--- %s ANTES1 ---" % (time.time() - start_time))

# % Util. produccion
for row in ws_util.iter_rows(min_row = 2, max_row = max_row_util, values_only=True):
  sector_oficina = row[0].lower()
  porcentaje = row[3]
  util_dicc[sector_oficina] = porcentaje
util_prod.close()
  
# Stock
for row in ws_stock.iter_rows(min_row = 2, max_row = max_row_stock, values_only=True):
  llave = row[0].lower()
  stock = row[3]
  dict_stock[llave] = stock
wb_stock.close()

# Por producir mes N
for row in ws_por_producir.iter_rows(min_row = 2, max_row = max_row_por_producir, values_only=True):
  dict_producir[row[1]] = row[2]

# Ponderación de cumplimiento
for value in ws_ponderaciones.iter_rows(min_row=2, max_row=5, values_only=True):
        dict_ponderacion_cumplimiento[value[0].lower()] = value[1]
por_producir.close()

# Producción por despachar mes N 
for row in ws_por_despachar.iter_rows(min_row = 2, max_row = max_row_por_despachar, values_only=True):
  llave = row[0].lower()
  kg_despachar = row[2]
  dict_por_despachar[llave] = kg_despachar
por_despachar.close()

# Delay 
for row in ws_delay.iter_rows(min_row = 2, max_row = max_row_delay, values_only=True):
  llave = row[0].lower()
  delay = row[3]
  dict_delay[llave] = delay
wb_delay.close()

# Vol. prom. Por contenedor
for row in ws_contenedor.iter_rows(min_row = 2, max_row = max_row_cont, values_only=True):
  llave = row[0].lower()
  volumen = row[3]
  dict_vol_cont[llave] = volumen
wb_vol_cont.close()

# En puerto a facturar
for row in ws_puerto.iter_rows(min_row = 2, max_row = max_row_puerto, values_only=True):
  llave = row[0].lower()
  puerto_total = row[2]
  dict_puerto[llave] = puerto_total
wb_puerto.close()

# Agregamos todo a la planilla 0.1 seg
for row in range(2, max_row + 1):
  llave_actual = ws_irrestricto.cell(row = row, column = 1).value.lower()         # Llave = 'Agro Mexico1012764'
  sector = ws_irrestricto.cell(row = row, column = 2).value                       # Sector = 'Pollo'
  oficina = ws_irrestricto.cell(row = row, column = 3).value                      # oficina = 'Agro Mexico'
  sku_actual = ws_irrestricto.cell(row = row, column = 4).value                   # sku = '1012764'
  
  concanate = (sector + oficina).lower()
  if concanate in util_dicc:
    ws_irrestricto.cell(row = row, column = 10).value = util_dicc[concanate]      # %Util.prod = '46%'
    ws_irrestricto['J'+ str(row)].number_format = FORMAT_PERCENTAGE
  
  elif concanate not in util_dicc:
    ws_irrestricto.cell(row = row, column = 10).value = 0.35
    ws_irrestricto['J' + str(row)].fill = PatternFill("solid", fgColor=red)

  if llave_actual in dict_stock:
    ws_irrestricto.cell(row = row, column = 12).value = dict_stock[llave_actual]
  
  elif llave_actual not in dict_stock:
    ws_irrestricto.cell(row = row, column = 12).value = 0 
    ws_irrestricto['L' + str(row)].fill = PatternFill("solid", fgColor=red)
  
  if sku_actual in dict_producir:
    ws_irrestricto.cell(row = row, column = 13).value = dict_producir[sku_actual] # Por producir mes N = "60.000"

  elif sku_actual not in dict_producir:
    ws_irrestricto.cell(row = row, column = 13).value = 0
    ws_irrestricto['M' + str(row)].fill = PatternFill("solid", fgColor=red)

  if llave_actual in dict_por_despachar:
    ws_irrestricto.cell(row = row, column = 14).value = dict_por_despachar[llave_actual]
  
  if llave_actual in dict_delay:
    ws_irrestricto.cell(row = row, column = 16).value = dict_delay[llave_actual]
  
  # !!!! Asumí que si no encuentro el delay en el excel, este tomará por defecto el valor 0
  elif llave_actual not in dict_delay:
    ws_irrestricto.cell(row = row, column = 16).value = 0

  if llave_actual in dict_vol_cont:
    ws_irrestricto.cell(row = row, column = 17).value = dict_vol_cont[llave_actual]
  
  elif llave_actual not in dict_vol_cont:
    ws_irrestricto.cell(row = row, column = 17).value = 24000
    ws_irrestricto['Q' + str(row)].fill = PatternFill("solid", fgColor=red)

  # En puerto a facturar
  if llave_actual in dict_puerto:
    ws_irrestricto.cell(row = row, column = 29).value = dict_puerto[llave_actual]
  
  # FORMATOS
  ws_irrestricto[f'K{row}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'L{row}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'M{row}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'N{row}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'O{row}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'P{row}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'Q{row}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'R{row}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'S{row}'].number_format = BUILTIN_FORMATS[3]

plan_irrestricto.save(filename)

# 4. Agregamos las columnas con formulas
# contador parte del 1
for i, row in enumerate(ws_irrestricto.iter_rows(min_row = 2, max_row = max_row, values_only=True), start = 2):
  # Prod. disponible
  prod = row[7] or 0                                            # Columna H
  util = row[9]                                                 # Columna J
  prod_disp = prod * util                                       # Columna K
  ws_irrestricto.cell(row = i, column = 11).value = prod_disp

  # Total disponible
  stock = row[11] or 0                                          # Columna L
  por_producir = row[12]                                        # Columna M
  produccion_por_despachar = row[13] or 0                       # Columna N
  sector = row[1].lower()                                       # Columna B
  ponderacion = dict_ponderacion_cumplimiento[sector]
  total_disp = (prod_disp + stock + por_producir * ponderacion) - produccion_por_despachar
  ws_irrestricto.cell(row = i, column = 15).value = total_disp  # Columna O                             

  # Atraso a facturar
  delay = row[15]                                               # Columna P
  vol_contenedor = row[16] or 1                                 # Columna Q 
  if total_disp >= delay and delay > 0:
    # atrasos = int(delay / vol_contenedor)
    # atraso_factu = atrasos * vol_contenedor
    # ws_irrestricto.cell(row = i, column = 18).value = atraso_factu
    ws_irrestricto.cell(row = i, column = 18).value = delay
  
  elif total_disp < delay and delay > 0:
    # ws_irrestricto.cell(row = i, column = 18).value = delay
    atrasos = int(delay / vol_contenedor)
    atraso_factu = atrasos * vol_contenedor
    ws_irrestricto.cell(row = i, column = 18).value = atraso_factu
  
  else:
    ws_irrestricto.cell(row = i, column = 18).value = 0

plan_irrestricto.save(filename)

for i, row in enumerate(ws_irrestricto.iter_rows(min_row = 2, max_row = max_row, values_only=True), start = 2):
  # Facturación atraso
  atraso_a_facturar = row[17]                                   # Columna R
  ajuste_atraso = row[18] or 0                                  # Columna S
  facturacion_atraso = atraso_a_facturar + ajuste_atraso
  ws_irrestricto.cell(row = i, column = 20).value = f'=R{i}-S{i}'
  
  # Producción para venta nueva
  total_disp = row[14]                                          # Columna O
  prd_venta_nueva = total_disp - facturacion_atraso
  ws_irrestricto.cell(row = i, column = 21).value = f'=O{i}-S{i}'

  # Venta del mes
  vol_contenedor = row[16] or 24000                             # Columna Q 
  venta_mes_n1 = row[8] or 0                                    # Columna J
  venta_mes = 0
  if prd_venta_nueva > vol_contenedor and venta_mes_n1 >= vol_contenedor:
    cant_contenedor = int(prd_venta_nueva / vol_contenedor)
    venta_mes = cant_contenedor * vol_contenedor                # Columna V
    ws_irrestricto.cell(row = i, column = 22).value = venta_mes

  else: ws_irrestricto.cell(row = i, column = 22).value = 0

  # Facturación Venta nueva                                      
  ajuste_venta_nueva = row[22] or 0                             # Columna W 
  facturacion_venta_nueva = 0
  if venta_mes > 0:
    ws_irrestricto.cell(row = i, column = 24).value = f'=V{i}+W{i}'
    facturacion_venta_nueva = venta_mes + ajuste_venta_nueva
  else: 
    ws_irrestricto.cell(row = i, column = 24).value = 0

  # Saldo Volumen disponible
  saldo_disp = prd_venta_nueva - facturacion_venta_nueva
  ws_irrestricto.cell(row = i, column = 25).value = f'=U{i}-X{i}'

  # Disponible stock sin venta
  delay = row[15]
  # print('\n', i, row[0], delay, saldo_disp, vol_contenedor, venta_mes_n1, total_disp)
  # print(saldo_disp)
  disponible_sin_venta = 0
  print(f'\n {i} {row[0]}: delay {delay} saldodisp {saldo_disp} vol {vol_contenedor} venta {venta_mes_n1} total {total_disp}')
  if saldo_disp > vol_contenedor and saldo_disp > venta_mes_n1 and delay < total_disp:
    cant_disp = int((saldo_disp + ajuste_venta_nueva) / vol_contenedor)
    disponible_sin_venta = cant_disp * vol_contenedor
    ws_irrestricto.cell(row = i, column = 26).value = disponible_sin_venta

  # Facturación stock
  ajuste_sin_venta = row[26] or 0                                   # Columna AA
  disponible_sin_venta = disponible_sin_venta or 0
  # print(ajuste_sin_venta, disponible_sin_venta)
  if disponible_sin_venta > 0:
    d_sin_venta = disponible_sin_venta + ajuste_sin_venta
    ws_irrestricto.cell(row = i, column = 28).value = f'=Z{i}+AA{i}'

  # Plan Irrestricto Inicial
  oficina = row[2]
  puerto_a_facturar = row[28] or 0
  if oficina == oficina_plan_irrestricto and delay > 0:
    ws_irrestricto.cell(row = i, column = 29).value = atraso_a_facturar
  
  elif oficina == oficina_plan_irrestricto and delay == 0:
    ws_irrestricto.cell(row = i, column = 29).value = puerto_a_facturar

  elif total_disp > 0:
    plan = atraso_a_facturar + venta_mes + disponible_sin_venta + puerto_a_facturar
    ws_irrestricto.cell(row = i, column = 29).value = plan

  # FORMATOS
  ws_irrestricto[f'T{i}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'U{i}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'V{i}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'W{i}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'X{i}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'Y{i}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'Z{i}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'AA{i}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'AB{i}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'AC{i}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'AD{i}'].number_format = BUILTIN_FORMATS[3]
  ws_irrestricto[f'AE{i}'].number_format = BUILTIN_FORMATS[3]


plan_irrestricto.save(filename)
plan_irrestricto.close()
print("--- %s seconds ---" % (time.time() - start_time))