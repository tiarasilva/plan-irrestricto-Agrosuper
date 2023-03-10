from openpyxl import load_workbook 
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE
from constants import *

from dateutil.relativedelta import relativedelta
from datetime import datetime, timedelta, date

def create_weighing_production(dict_tipo_venta, filename_util, selected_month, selected_year):
  wb_fecha_zarpe = load_workbook(filename_util)
  ws_datos_faena = wb_fecha_zarpe['Datos Faena y Días']

  if 'Ponderación' in wb_fecha_zarpe.sheetnames:
    del wb_fecha_zarpe['Ponderación']
  ws_ponderacion = wb_fecha_zarpe.create_sheet('Ponderación')
  ws_ponderacion.append(['Año', 'Mes', 'Llave', 'Sector', 'Oficina', 'Corte de producción', 'Ponderación', 'Fechas'])

  # ------ Parametros ------
  number_selected_month = month_number[selected_month.lower()]

  month_1 = date(int(selected_year), int(number_selected_month), 1)
  month_2 = month_1 + relativedelta(months=1)
  month_3 = month_1 + relativedelta(months=2)

  dict_sector_ponderacion = {}
  name_month_1 = month_1.strftime('%B').lower()
  name_month_2 = month_2.strftime('%B').lower()
  name_month_3 = month_3.strftime('%B').lower()
  
  # ------ Agregamos las columnas fijas de Venta Directa
  j = 2
  for time in [month_1, month_2, month_3]:
    month = time.strftime('%B')
    key = f'{time.year}{month.lower()}'

    if key not in dict_sector_ponderacion:
      dict_sector_ponderacion[key] = {}

    for sector in ['Cerdo', 'Pollo', 'Pavo', 'Elaborado']:
      for oficina in dict_tipo_venta[seleccion_tipo_venta]:
        ws_ponderacion.cell(row = j, column = 1).value = time.year
        ws_ponderacion.cell(row = j, column = 2).value = month_translate_EN_CL[month]
        ws_ponderacion.cell(row = j, column = 3).value = sector + oficina
        ws_ponderacion.cell(row = j, column = 4).value = sector
        ws_ponderacion.cell(row = j, column = 5).value = oficina
        dict_sector_ponderacion[key][f'{sector.lower()}{oficina.lower()}'] = 0
        j += 1

  # ------ Leemos la data ------
  dict_ponderacion_total_mensual = {}
  total_cerdo = 0
  total_pollo = 0

  for row in ws_datos_faena.iter_rows(3, ws_datos_faena.max_row, values_only=True):
    if row[3] is None:
      break
    process_tag_cerdo = row[0]
    process_tag_pollo = row[7]
    date_row = row[3]
    month = date_row.strftime('%B')
    weight_cerdo = row[4]
    weight_pollo = row[11]
    key = f'{date_row.year}{month.lower()}'

    if key in dict_ponderacion_total_mensual:
      dict_ponderacion_total_mensual[key] += weight_cerdo

    else:
      dict_ponderacion_total_mensual[key] = weight_cerdo
      total_cerdo = 0
      total_pollo = 0

    # cerdo
    if key in dict_sector_ponderacion:
      if process_tag_cerdo is None:
        total_cerdo += weight_cerdo
      
      elif process_tag_cerdo is not None:
        if 'Producción' in process_tag_cerdo:
          if 'Stacking' in process_tag_cerdo or 'Zarpe' in process_tag_cerdo:
            lista_cerdo = process_tag_cerdo.split('\n')
            len_lista_cerdo = len(lista_cerdo) - 1
            process_tag_cerdo = lista_cerdo[len_lista_cerdo]
          tipo_tag = process_tag_cerdo.split(': ')
          lista_tag = tipo_tag[1].split(', ')
          total_cerdo += weight_cerdo
          
          for item in lista_tag:
            second_key = 'cerdo' + item.lower()
            if second_key in dict_sector_ponderacion[key]:
              dict_sector_ponderacion[key][second_key] = total_cerdo
      
      # pollo
      if process_tag_pollo is None:
        total_pollo += weight_pollo
      
      elif process_tag_pollo is not None:
        if 'Producción' in process_tag_pollo:
          if 'Stacking' in process_tag_pollo or 'Zarpe' in process_tag_pollo:
            lista_pollo = process_tag_pollo.split('\n')
            len_lista_pollo = len(lista_pollo) - 1
            process_tag_pollo = lista_pollo[len_lista_pollo]
          tipo_tag_pollo = process_tag_pollo.split(': ')
          lista_tag_pollo = tipo_tag_pollo[1].split(', ')
          total_pollo += weight_pollo
          
          for oficina in lista_tag_pollo:
            second_key = 'pollo' + oficina.lower()
            s_k_pavo = 'pavo' + oficina.lower()
            s_k_elab = 'elaborado' + oficina.lower()
            if second_key in dict_sector_ponderacion[key]:
              dict_sector_ponderacion[key][second_key] = total_pollo
            if s_k_pavo in dict_sector_ponderacion[key]:
              dict_sector_ponderacion[key][s_k_pavo] = total_pollo
            if s_k_elab in dict_sector_ponderacion[key]:
              dict_sector_ponderacion[key][s_k_elab] = total_pollo

  for k, r in enumerate(ws_ponderacion.iter_rows(2, ws_ponderacion.max_row, values_only = True), start = 2):
    year = 0
    month = ''

    if r[2] is None:
      break

    if r[0] is not None:
      year = r[0]

    if r[1] is not None:
      month = r[1].lower()
      EN_month = month_translate_CL_EN[month]

    EN_month = month_translate_CL_EN[month]
    first_key = f'{year}{EN_month}'
    second_key = r[2]
    second_key = second_key.lower()

    pon_1 = dict_sector_ponderacion[first_key][second_key]
    pon_tot = dict_ponderacion_total_mensual[first_key]

    ws_ponderacion[f'F{k}'].value = pon_1
    ws_ponderacion[f'F{k}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws_ponderacion[f'G{k}'].value = pon_1 / pon_tot
    ws_ponderacion[f'G{k}'].number_format = FORMAT_PERCENTAGE
  
  ultimate_max = ws_ponderacion.max_row + 1
  ws_ponderacion[f'F{ultimate_max}'].value = 'Ponderación promedio' 
  ws_ponderacion[f'G{ultimate_max}'] = f'=AVERAGE(G2:G{ultimate_max - 1})'
  ws_ponderacion[f'G{ultimate_max}'].number_format = FORMAT_PERCENTAGE

  for i in range(1, 8):
    ws_ponderacion[f'{get_column_letter(i)}{ultimate_max}'].fill = PatternFill("solid", fgColor=blue)
    ws_ponderacion[f'{get_column_letter(i)}{ultimate_max}'].font = Font(bold=True, color=white)
  

  # ------ Estilos ------
  for i in range(1, 9):
    thin = Side(border_style="thin", color=white)
    ws_ponderacion[f'{get_column_letter(i)}1'].font = Font(bold=True, color=white)  # tomamos la primera fila
    ws_ponderacion[f'{get_column_letter(i)}1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_ponderacion[f'{get_column_letter(i)}1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws_ponderacion[f'{get_column_letter(i)}1'].fill = PatternFill("solid", fgColor=lightBlue)
  
  ws_ponderacion.column_dimensions['B'].width = 12    # Mes
  ws_ponderacion.column_dimensions['C'].width = 25    # Llave
  ws_ponderacion.column_dimensions['D'].width = 12    # Sector
  ws_ponderacion.column_dimensions['E'].width = 20    # Oficina
  ws_ponderacion.column_dimensions['F'].width = 20

  # Merge cells
  ws_ponderacion.merge_cells('A2:A21')
  ws_ponderacion.merge_cells('A22:A41')
  ws_ponderacion.merge_cells('A42:A61')

  ws_ponderacion.merge_cells('B2:B21')
  ws_ponderacion.merge_cells('B22:B41')
  ws_ponderacion.merge_cells('B42:B61')

  ws_ponderacion['A2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws_ponderacion['A22'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws_ponderacion['A42'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

  ws_ponderacion['B2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws_ponderacion['B22'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws_ponderacion['B42'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

  # Linea separadora azul
  line_blue = Side(border_style="thin", color=blue)

  for col in range(1, 8):
    ws_ponderacion[f'{get_column_letter(col)}21'].border = Border(bottom=line_blue)
    ws_ponderacion[f'{get_column_letter(col)}41'].border = Border(bottom=line_blue)
    ws_ponderacion[f'{get_column_letter(col)}61'].border = Border(bottom=line_blue)
  
  # print(dict_ponderacion_total_mensual, '\n')
  # print(dict_sector_ponderacion)
  wb_fecha_zarpe.save(filename_util)
  wb_fecha_zarpe.close()