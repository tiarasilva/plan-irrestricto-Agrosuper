# Fechas de zarpe - Logistica -> % Util. produccion
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from constants import *
from datetime import datetime, timedelta, date
from dateutil import rrule
import holidays
import calendar

def fechas_zarpe(dict_tipo_venta, filename_parametros, filename_util):
  ## ------ PARAMETROS ------
  wb_parametro = load_workbook(filename_parametros, data_only=True, read_only=True)
  ws_cong_consol = wb_parametro['Congelado y consolidación']
  dict_tiempo_consolidado = {}

  for row in ws_cong_consol.iter_rows(2, ws_cong_consol.max_row, values_only = True):
    if row[0] is None:
      break
    oficina = row[0]
    sector = row[1]
    total = row[4]
    llave = oficina.lower() + sector.lower()
    dict_tiempo_consolidado[llave] = total
  wb_parametro.close()

  ## ------ CALENDARIO ZARPE ------
  wb_fecha_zarpe = load_workbook(filename_util)
  ws_fecha_zarpe = wb_fecha_zarpe['Ultima fecha de Zarpe']
  util_dicc = {}
  dict_stacking = {}
  dict_zarpe = {}

  for row in ws_fecha_zarpe.iter_rows(min_row = 4, max_row = ws_fecha_zarpe.max_row, values_only=True):
    if row[1] is None:
      break
    mes = row[1]
    oficina = row[2].strip()
    item = row[3]
    fecha = datetime.date(row[4])
    llave = mes.lower() + oficina.lower()

    if oficina.lower() in dict_tipo_venta[seleccion_tipo_venta]:
      if 'stacking' in item.lower():
        if fecha in dict_stacking:
          dict_stacking[fecha].append(oficina)
        else:
          dict_stacking[fecha] = [oficina]
      
      if 'zarpe' in item.lower():
        if fecha in dict_zarpe:
          dict_zarpe[fecha].append(oficina)
        else:
          dict_zarpe[fecha] = [oficina]

  if 'Datos Faena y Días' in wb_fecha_zarpe.sheetnames:
    del wb_fecha_zarpe['Datos Faena y Días']
  ws_datos_faena = wb_fecha_zarpe.create_sheet()
  ws_datos_faena.title = 'Datos Faena y Días'

  ws_datos_faena.append({2: 'Cerdo', 9: 'Pollo'})

  # Fecha corte de producción
  ws_datos_faena.append({2: 'Mes', 3: 'Día', 4: 'Fecha', 5: 'Ponderación', 6: 'Ponderación mensual', 9: 'Mes', 10: 'Día', 11: 'Fecha', 12: 'Ponderación', 13:'Ponderación mensual'})
  thin = Side(border_style="thin", color=white)
  ws_datos_faena['B1'].font = Font(bold=True, color=white)  # tomamos la primera fila
  ws_datos_faena['B1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws_datos_faena['B1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws_datos_faena['B1'].fill = PatternFill("solid", fgColor=blue)

  ws_datos_faena['I1'].font = Font(bold=True, color=white)  # tomamos la primera fila
  ws_datos_faena['I1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
  ws_datos_faena['I1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
  ws_datos_faena['I1'].fill = PatternFill("solid", fgColor=blue)

  for i in range(2, 7):
    ws_datos_faena[f'{get_column_letter(i)}2'].font = Font(bold=True, color=white)  # tomamos la primera fila
    ws_datos_faena[f'{get_column_letter(i)}2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_datos_faena[f'{get_column_letter(i)}2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws_datos_faena[f'{get_column_letter(i)}2'].fill = PatternFill("solid", fgColor=lightBlue)

    ws_datos_faena[f'{get_column_letter(i + 7)}2'].font = Font(bold=True, color=white)  # tomamos la primera fila
    ws_datos_faena[f'{get_column_letter(i + 7)}2'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws_datos_faena[f'{get_column_letter(i + 7)}2'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws_datos_faena[f'{get_column_letter(i + 7)}2'].fill = PatternFill("solid", fgColor=lightBlue)

  # Agregamos fecha de inicio al calendario
  obj = calendar.Calendar()
  i = 2
  dict_row_prod = {'cerdo': 'E', 'pollo': 'L'}
  dict_row_descr = {'cerdo': 'A', 'pollo': 'H'}

  data_inicio = ws_fecha_zarpe['E4'].value
  mes_inicio = data_inicio.month
  año_inicio = data_inicio.year

  start_date = datetime(año_inicio, mes_inicio, 1)
  end_date = datetime(año_inicio + 1, mes_inicio, 1)

  for dt in rrule.rrule(rrule.MONTHLY, dtstart=start_date, until=end_date):
    res = calendar.monthrange(dt.year, dt.month)
    last_day = res[1]
    date_last_day = date(dt.year, dt.month, last_day)

    for day in obj.itermonthdates(dt.year, dt.month):
      if day.month == dt.month:
        i += 1

        if day.strftime('%A') == 'Sunday' and day not in holidays.CL(years=year):
          ws_datos_faena.cell(row = i, column = 5).value = 0
          ws_datos_faena.cell(row = i, column = 12).value = 0

        elif day.strftime('%A') == 'Saturday' and day not in holidays.CL(years=year):
          ws_datos_faena.cell(row = i, column = 5).value = 0.33
          ws_datos_faena.cell(row = i, column = 12).value = 0.33

        elif day in holidays.CL(years=year):
          ws_datos_faena.cell(row = i, column = 5).value = 0
          ws_datos_faena.cell(row = i, column = 12).value = 0
          ws_datos_faena[f'C{i}'].fill = PatternFill("solid", fgColor=lightOrange)
          ws_datos_faena[f'D{i}'].fill = PatternFill("solid", fgColor=lightOrange)
          ws_datos_faena[f'E{i}'].fill = PatternFill("solid", fgColor=lightOrange)

          ws_datos_faena[f'J{i}'].fill = PatternFill("solid", fgColor=lightOrange)
          ws_datos_faena[f'K{i}'].fill = PatternFill("solid", fgColor=lightOrange)
          ws_datos_faena[f'L{i}'].fill = PatternFill("solid", fgColor=lightOrange)

          if (day - timedelta(1)) not in holidays.CL(years=year):
            ws_datos_faena.cell(row = i-1, column = 5).value = 0.67
            ws_datos_faena.cell(row = i-1, column = 12).value = 0.67
            ws_datos_faena[f'C{i-1}'].fill = PatternFill("solid", fgColor=lightPale)
            ws_datos_faena[f'D{i-1}'].fill = PatternFill("solid", fgColor=lightPale)
            ws_datos_faena[f'E{i-1}'].fill = PatternFill("solid", fgColor=lightPale)

            ws_datos_faena[f'J{i-1}'].fill = PatternFill("solid", fgColor=lightPale)
            ws_datos_faena[f'K{i-1}'].fill = PatternFill("solid", fgColor=lightPale)
            ws_datos_faena[f'L{i-1}'].fill = PatternFill("solid", fgColor=lightPale)

        else:
          ws_datos_faena.cell(row = i, column = 5).value = 1
          ws_datos_faena.cell(row = i, column = 12).value = 1
        
        # Stacking
        if day in dict_stacking:
          for item in dict_stacking[day]:
            valor_linea_stacking = ws_datos_faena[f'A{i}'].value
            if valor_linea_stacking is None:
              ws_datos_faena.cell(row = i, column = 1).value = f'Stacking: {item}'
              ws_datos_faena.cell(row = i, column = 8).value = f'Stacking: {item}'
            else:
              ws_datos_faena.cell(row = i, column = 1).value = f'{valor_linea_stacking}, {item}'
              ws_datos_faena.cell(row = i, column = 8).value = f'{valor_linea_stacking}, {item}'
            ws_datos_faena[f'A{i}'].fill = PatternFill("solid", fgColor=yellow)
            ws_datos_faena[f'H{i}'].fill = PatternFill("solid", fgColor=yellow)
            ws_datos_faena[f'A{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws_datos_faena[f'H{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws_datos_faena[f'A{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            ws_datos_faena[f'H{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
            
            # Producción:
            for sector in ['cerdo', 'pollo']:
              tiempo_consolidado = dict_tiempo_consolidado[f'{item.lower()}{sector}']
              letter = dict_row_prod[sector]
              l_prod = dict_row_descr[sector]
              t_mayor = i
              suma_mayor = ws_datos_faena[f'{letter}{t_mayor}'].value               # E211

              # Buscamos una ponderación que supere el tiempo consolidado
              while suma_mayor <= tiempo_consolidado:
                t_mayor -= 1 
                suma_mayor += ws_datos_faena[f'{letter}{t_mayor}'].value             # E199

              # Buscamos el siguiente menor al maximo tiempo de consolidado
              t_menor = t_mayor + 1                                                  # E200
              suma_menor = suma_mayor - ws_datos_faena[f'{letter}{t_mayor}'].value   # SUMA(E198:E211)
              while suma_menor >= suma_mayor:
                suma_menor -= ws_datos_faena[f'{letter}{t_menor}'].value             # t_menor = E198 --> SUMA(E197:E211)
                t_menor += 1                                                         # E197

              # Chequeamos cual tiempo esta más cercano al tiempo consolidado:
              t_prod = t_menor
              if abs(suma_mayor - tiempo_consolidado) < abs(suma_menor - tiempo_consolidado):
                t_prod = t_mayor
              
              valor_linea = ws_datos_faena[f'{l_prod}{t_prod}'].value
              if valor_linea is None:
                ws_datos_faena[f'{l_prod}{t_prod}'].value = f'Producción: {item}'
              else:
                if 'Stacking' in valor_linea or 'Zarpe' in valor_linea:
                  if 'Producción' not in valor_linea:
                    ws_datos_faena[f'{l_prod}{t_prod}'].value = f'\nProducción: {oficina}'
                ws_datos_faena[f'{l_prod}{t_prod}'].value = f'{valor_linea}, {item}'
                
              # Estilos
              ws_datos_faena[f'{l_prod}{t_prod}'].fill = PatternFill("solid", fgColor=grey)
              ws_datos_faena[f'{l_prod}{t_prod}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
              ws_datos_faena[f'{l_prod}{t_prod}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                
        # Zarpe
        if day in dict_zarpe:
          office_zarpe = ''
          for of in dict_zarpe[day]:
            office_zarpe += ', ' + str(of)
          ws_datos_faena.cell(row = i, column = 1).value = f'Zarpe: {office_zarpe}'
          ws_datos_faena.cell(row = i, column = 8).value = f'Zarpe: {office_zarpe}'
          ws_datos_faena[f'A{i}'].fill = PatternFill("solid", fgColor=lightGreen)
          ws_datos_faena[f'H{i}'].fill = PatternFill("solid", fgColor=lightGreen)
          ws_datos_faena[f'A{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
          ws_datos_faena[f'H{i}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
          ws_datos_faena[f'A{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
          ws_datos_faena[f'H{i}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)

        if day == date_last_day:
          for sector in ['cerdo', 'pollo']:
            for oficina in ['Agrosuper Brasil', 'Exportacion Directa']:
              tiempo_consolidado = dict_tiempo_consolidado[f'{oficina.lower()}{sector}']
              t_menor = int(tiempo_consolidado)
              t_mayor = 0

              suma_mayor = ws_datos_faena[f'{dict_row_prod[sector]}{i - t_mayor}'].value
              while suma_mayor <= tiempo_consolidado:
                t_mayor += 1
                suma_mayor += ws_datos_faena[f'{dict_row_prod[sector]}{i - t_mayor}'].value
              suma_menor = suma_mayor - ws_datos_faena[f'{dict_row_prod[sector]}{i - t_mayor}'].value
              t_menor = t_mayor - 1

              t_prod = t_menor
              if abs(suma_mayor - tiempo_consolidado) < abs(suma_menor - tiempo_consolidado):
                t_prod = t_mayor

              valor_linea = ws_datos_faena[f'{dict_row_descr[sector]}{i - t_prod}'].value
              if valor_linea is None:
                ws_datos_faena[f'{dict_row_descr[sector]}{i - t_prod}'].value = f'Producción: {oficina}'
              else:
                if 'Stacking' in valor_linea or 'Zarpe' in valor_linea:
                  if 'Producción' not in valor_linea:
                    ws_datos_faena[f'{dict_row_descr[sector]}{i - t_prod}'].value = f'{valor_linea} \nProducción: {oficina}'
                else:
                  ws_datos_faena[f'{dict_row_descr[sector]}{i - t_prod}'].value = f'{valor_linea}, {oficina}'

              ws_datos_faena[f'{dict_row_descr[sector]}{i - t_prod}'].fill = PatternFill("solid", fgColor=grey)
              ws_datos_faena[f'{dict_row_descr[sector]}{i - t_prod}'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
              ws_datos_faena[f'{dict_row_descr[sector]}{i - t_prod}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
          
          # Suma total
          len_month = calendar.monthrange(day.year, day.month)[1] - 1
          inicio_mes = i - len_month
          ws_datos_faena.merge_cells(f'F{inicio_mes}:F{i}')
          ws_datos_faena[f'F{inicio_mes}'].value = f'=SUM(E{inicio_mes}:E{i})'
          ws_datos_faena[f'F{inicio_mes}'].alignment = Alignment(horizontal="center", vertical="center")

          ws_datos_faena.merge_cells(f'M{inicio_mes}:M{i}')
          ws_datos_faena[f'M{inicio_mes}'].value = f'=SUM(L{inicio_mes}:L{i})'
          ws_datos_faena[f'M{inicio_mes}'].alignment = Alignment(horizontal="center", vertical="center")

          # Nombre del mes
          month_EN = day.strftime('%B')
          ws_datos_faena.merge_cells(f'B{inicio_mes}:B{i}')
          ws_datos_faena[f'B{inicio_mes}'].value = month_translate_EN_CL[month_EN]
          ws_datos_faena[f'B{inicio_mes}'].alignment = Alignment(horizontal="center", vertical="center")

          ws_datos_faena.merge_cells(f'I{inicio_mes}:I{i}')
          ws_datos_faena[f'I{inicio_mes}'].value = month_translate_EN_CL[month_EN]
          ws_datos_faena[f'I{inicio_mes}'].alignment = Alignment(horizontal="center", vertical="center")
        
        week_EN = day.strftime('%A')
        ws_datos_faena[f'C{i}'].value = week_translate_EN_CL[week_EN]
        ws_datos_faena[f'D{i}'].value = day
        ws_datos_faena[f'J{i}'].value = week_translate_EN_CL[week_EN]
        ws_datos_faena[f'K{i}'].value = day
  wb_fecha_zarpe.save(filename_util)

  ws_datos_faena.column_dimensions['A'].width = 25
  ws_datos_faena.column_dimensions['B'].width = 10
  ws_datos_faena.column_dimensions['C'].width = 12
  ws_datos_faena.column_dimensions['D'].width = 12
  ws_datos_faena.column_dimensions['F'].width = 15

  ws_datos_faena.column_dimensions['H'].width = 25
  ws_datos_faena.column_dimensions['I'].width = 10
  ws_datos_faena.column_dimensions['J'].width = 12
  ws_datos_faena.column_dimensions['K'].width = 12
  ws_datos_faena.column_dimensions['M'].width = 15

  ws_datos_faena['O3'].fill = PatternFill("solid", fgColor=lightOrange)
  ws_datos_faena.cell(row = 3, column = 15).value = 0
  ws_datos_faena.cell(row = 3, column = 16).value = 'Feriado'

  ws_datos_faena['O4'].fill = PatternFill("solid", fgColor=lightPale)
  ws_datos_faena.cell(row = 4, column = 15).value = 0.67
  ws_datos_faena.cell(row = 4, column = 16).value = 'Preferiado'

  ws_datos_faena.cell(row = 5, column = 15).value = 0.33
  ws_datos_faena.cell(row = 5, column = 16).value = 'Sábado'

  ws_datos_faena['O6'].fill = PatternFill("solid", fgColor=yellow)
  ws_datos_faena.cell(row = 6, column = 16).value = 'Stacking'

  ws_datos_faena['O7'].fill = PatternFill("solid", fgColor=grey)
  ws_datos_faena.cell(row = 7, column = 16).value = 'Producción'

  wb_fecha_zarpe.save(filename_util)
  wb_fecha_zarpe.close()