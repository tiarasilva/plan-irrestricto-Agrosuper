from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from constants import *
from styles import run_styles
from message import messageBox
from createWeighing import create_weighing_production
from fechasZarpe import fechas_zarpe
import time

import sys, os
from os import path
import pathlib

start_time = time.time()

if getattr(sys, 'frozen', False):
  print('\nRunning in a PyInstaller bundle\n')
  bundle_dir = sys._MEIPASS
  filename_asignacion = path.abspath(path.join(path.dirname(__file__), filename_asignacion))
  filename_maestro_materiales = path.abspath(path.join(path.dirname(__file__), filename_maestro_materiales))
  filename_parametros = path.abspath(path.join(path.dirname(__file__), filename_parametros))
  filename_util = path.abspath(path.join(path.dirname(__file__), filename_util))
  filename_stock = path.abspath(path.join(path.dirname(__file__), filename_stock))
  filename_por_producir = path.abspath(path.join(path.dirname(__file__), filename_por_producir))
  filename_por_despachar = path.abspath(path.join(path.dirname(__file__), filename_por_despachar))
  filename_vol_cont = path.abspath(path.join(path.dirname(__file__), filename_vol_cont))
  filename_puerto = path.abspath(path.join(path.dirname(__file__), filename_puerto))
  path_img = path.abspath(path.join(path.dirname(__file__), path_img))
  filename = path.abspath(path.join(path.dirname(sys.executable), filename))
  
else:
  print('\nRunning in a Python environment\n')
  bundle_dir = os.path.dirname(os.path.abspath(__file__))

# 1. Crear new Workbook/Excel
wb = Workbook()
ws = wb.active
ws.title = "BASE PLAN MES N"
max_row = ws.max_row
max_col = ws.max_column

# 2. Agregamos los titulos de las columnas
ws.append(['Llave',
           'Sector',                        
           'Oficina',                       
           'Material',
           'Descripción',
           'Nivel Jer. 2',
           'Nivel Jer. 3',
           'RV Producción mes n+1',
           'RV Venta mes n+1',
           '% Uti. producción',
           'Producción disponible',
           'Stock al día',
           'Por producir mes N',
           'Producción por despachar mes N',
           'Total disponible',
           'Delay',
           'Vol. prom. Por contenedor',
           'Atraso a facturar',
           'Ajuste atraso',
           'Facturación atraso',
           'Producción para venta nueva',
           'Venta del mes',
           'Ajuste venta nueva',
           'Facturación Venta nueva',
           'Saldo Volumen disponible',
           'Disponible stock sin venta',
           'Ajuste stock sin venta',
           'Facturación stock',
           'En puerto a facturar',
           'Plan Irrestricto',
           'Plan Ajustado',
           'Motivo Ajuste'
           ])

# 3. Agregamos la data
asignacion = load_workbook(filename_asignacion, data_only=True, read_only=True)
ws_produccion = asignacion['RV Producción']
max_row_new = ws.max_row
max_rows = max_row_new
dict_produccion_inicial = {}

# 3.1 Maestro de materiales
wb_maestro_materiales = load_workbook(filename_maestro_materiales, data_only=True, read_only=True)
ws_maestro_materiales = wb_maestro_materiales.active
max_maestro = ws_maestro_materiales.max_row
dict_materiales = {}

for row in ws_maestro_materiales.iter_rows(min_row = 2, max_row = max_maestro, values_only=True):
  material = row[1]
  descripcion = row[2]
  sector = row[3]
  nivel_2 = row[4]
  nivel_3 = row[5]
  estado = row[6]
  mercado = row[7]
  vida_util = row[8]
  dict_materiales[material] = {'descripcion': descripcion, 
                               'sector': sector,
                               'nivel_2': nivel_2, 
                               'nivel_3': nivel_3, 
                               'estado': estado,
                               'mercado': mercado,
                               'vida_util': vida_util
                               }
wb_maestro_materiales.close()

# 3.2. Parametros
## ------ Chequear si esta en VENTA DIRECTA
wb_parametros = load_workbook(filename_parametros, data_only = True, read_only = True)
ws_parametros_venta = wb_parametros['Tipo de venta']
ws_ponderaciones = wb_parametros['Ponderación cumplimiento produc']
ws_parametros = wb_parametros['Parametros']

parametro_porcentaje = ws_parametros['B1'].value
month_of_projection = ws_parametros['B2'].value
selected_year = ws_parametros['B3'].value
dict_tipo_venta = {"Venta Directa": [], "Venta Local": []}
tipo_venta = "Venta Directa"
dict_ponderacion_cumplimiento = {}

for row in ws_parametros_venta.iter_rows(min_row = 2, max_row = ws_parametros_venta.max_row, values_only = True):
  if row[1] is None:
    break
  if row[0] == "Venta Local":
    tipo_venta = row[0]
  dict_tipo_venta[tipo_venta].append(row[1].lower())

# ----- Ponderación de cumplimiento
for value in ws_ponderaciones.iter_rows(min_row=2, max_row=5, values_only=True):
  dict_ponderacion_cumplimiento[value[0].lower()] = value[1]
wb_parametros.close()

# 3.3. PRODUCCION
dict_otro_tipo = {}
i = 2
for row in ws_produccion.iter_rows(min_row = 2, max_row = ws_produccion.max_row, values_only = True):
  if row[0] is None:
    break
  llave = row[0]
  sku = row[1]
  oficina = row[2]
  descripcion = row[3]
  prd_mes = row[4]
  dict_produccion_inicial[llave.lower()] = prd_mes

  if oficina.lower() in dict_tipo_venta[seleccion_tipo_venta] and prd_mes is not None:
    nivel_2 = dict_materiales[str(sku)]['nivel_2']
    if oficina.lower() == 'agro mexico':
      if nivel_2 == 'Carne Recuperada':
        ws.cell(row = i, column = 1).value = llave         # A: LLave
        ws.cell(row = i, column = 3).value = oficina       # C: Oficina
        ws.cell(row = i, column = 4).value = sku           # D: Material
        ws.cell(row = i, column = 5).value = descripcion   # E: Descripcion
        ws.cell(row = i, column = 8).value = prd_mes       # H: RV Producción mes n+1
    else:
      ws.cell(row = i, column = 1).value = llave         # A: LLave
      ws.cell(row = i, column = 3).value = oficina       # C: Oficina
      ws.cell(row = i, column = 4).value = sku           # D: Material
      ws.cell(row = i, column = 5).value = descripcion   # E: Descripcion
      ws.cell(row = i, column = 8).value = prd_mes       # H: RV Producción mes n+1

    if str(sku) in dict_materiales.keys():
      ws[f'B{i}'].value = dict_materiales[str(sku)]['sector']
      ws[f'E{i}'].value = dict_materiales[str(sku)]['descripcion'] # E:Descripción
      ws[f'F{i}'].value = dict_materiales[str(sku)]['nivel_2']     # F: 
      ws[f'G{i}'].value = dict_materiales[str(sku)]['nivel_3']     # G:
    
    else:
      if 'PV' in descripcion:
        ws[f'E{i}'].value = PV      
      elif 'PO' in descripcion:
        ws[f'E{i}'].value = PO                              
      elif 'GO' in descripcion:
        ws[f'E{i}'].value = GO    
      elif 'GA' in descripcion:
        ws[f'E{i}'].value = GA
      else:
        ws[f'E{i}'].value = 'Elaborado' # Sector = 'Elaborado'
    i += 1
  
  # Si es venta local
  else:
    dict_otro_tipo[llave] = {'sku': sku, 'Oficina': oficina, 'Descripción': descripcion, 'Producción mes': prd_mes}
wb.save(filename)
wb.close()

# ----- VENTA MES n+1 
plan_irrestricto = load_workbook(filename)
ws_irrestricto = plan_irrestricto.active
max_row = ws_irrestricto.max_row

venta = asignacion['RV Plan de ventas']
dict_venta_n1 = {}
dict_venta_no_asignadas = {}

for row in venta.iter_rows(min_row = 2, max_row = venta.max_row, values_only=True):
  if row[0] is None:
    break
  key = row[0].lower()
  office = row[1]
  sku = row[2]
  description = row[3]
  monthly_sales = row[4]
  dict_venta_n1[key] = monthly_sales
  dict_venta_no_asignadas[key] = {'office': office, 'sku': sku, 'description': description, 'monthly_sales': monthly_sales}
asignacion.close()

for i, row in enumerate(ws_irrestricto.iter_rows(min_row = 2, max_row = ws_irrestricto.max_row, values_only=True), start = 2):
  llave_actual = row[0]
  if llave_actual is None:
    break
  
  if llave_actual.lower() in dict_venta_n1:
    venta = dict_venta_n1[llave_actual.lower()]
    if venta >= 0:
      dict_venta_no_asignadas.pop(llave_actual.lower(), None)
      ws_irrestricto[f'I{i}'].value = venta
    else:
      ws_irrestricto[f'I{i}'].value = 0
      ws_irrestricto[f'I{i}'].fill = PatternFill("solid", fgColor=orange)
  
  else:
    ws_irrestricto[f'I{i}'].value = 0

# ----- Ventas sin producción
j = ws_irrestricto.max_row
for key, value in dict_venta_no_asignadas.items():
  if value['office'].lower() in dict_tipo_venta[seleccion_tipo_venta]:
    if value['office'].lower() == 'agro mexico':
      nivel_2 = dict_materiales[str(value['sku'])]['nivel_2']
      if nivel_2.lower() == 'carne recuperada':
        j += 1
        ws_irrestricto[f'A{j}'].value = key
        ws_irrestricto[f'C{j}'].value = value['office']
        ws_irrestricto[f'D{j}'].value = value['sku']
        ws_irrestricto[f'E{j}'].value = value['description']
        ws_irrestricto[f'H{j}'].value = 0
    else:
        j += 1
        ws_irrestricto[f'A{j}'].value = key
        ws_irrestricto[f'C{j}'].value = value['office']
        ws_irrestricto[f'D{j}'].value = value['sku']
        ws_irrestricto[f'E{j}'].value = value['description']
        ws_irrestricto[f'H{j}'].value = 0
    
    venta = value['monthly_sales'] or 0
    if venta >= 0:
      ws_irrestricto[f'I{j}'].value = venta
    else:
      ws_irrestricto[f'I{j}'].value = 0
      ws_irrestricto[f'I{j}'].fill = PatternFill("solid", fgColor=orange)

    if str(value['sku']) in dict_materiales.keys():
      ws_irrestricto[f'B{j}'].value = dict_materiales[str(value['sku'])]['sector']
      ws_irrestricto[f'E{j}'].value = dict_materiales[str(value['sku'])]['descripcion']
      ws_irrestricto[f'F{j}'].value = dict_materiales[str(value['sku'])]['nivel_2']
      ws_irrestricto[f'G{j}'].value = dict_materiales[str(value['sku'])]['nivel_3']
    
    else:
      if 'PV' in value['description']:
        ws_irrestricto[f'B{j}'].value = PV      
      elif 'PO' in value['description']:
        ws_irrestricto[f'B{j}'].value = PO                              
      elif 'GO' in value['description']:
        _ws_irrestricto[f'B{j}'].value = GO    
      elif 'GA' in value['description']:
        ws_irrestricto[f'B{j}'].value = GA
      else:
        ws_irrestricto[f'B{j}'].value = 'Elaborado' # Sector = 'Elaborado'
plan_irrestricto.save(filename)

# Fechas de zarpe - Logistica -> % Util. produccion
fechas_zarpe(dict_tipo_venta, filename_parametros, filename_util)
create_weighing_production(dict_tipo_venta, filename_util, month_of_projection, selected_year)

util_prod = load_workbook(filename_util, data_only=True)
ws_util = util_prod['Ponderación']
util_dicc = {}
year = ''
month = ''

for row in ws_util.iter_rows(min_row = 2, max_row = ws_util.max_row, values_only=True):
  if row[0] is not None:
    year = row[0]
  
  if row[1] is not None:
    month = row[1]

  if row[2] is None:
    break

  sector_oficina = row[2].lower()
  porcentaje = row[5]
  EN_month = month_translate_CL_EN[month.lower()]
  selected_month = month_translate_CL_EN[month_of_projection.lower()]

  if EN_month == selected_month.lower() and int(year) == int(selected_year):
    util_dicc[sector_oficina] = porcentaje
util_prod.close()

# Stock
wb_stock = load_workbook(filename_stock, data_only = True, read_only=True)
ws_stock = wb_stock['TD Stock']
ws_delay = wb_stock['DELAY']
dict_delay = {}
dict_stock = {}
dict_stock_no_asignado = {}

for row in ws_stock.iter_rows(min_row = 2, max_row = ws_stock.max_row - 1, values_only=True):
  if row[0] is None:
    break
  sku = row[0]
  oficina = row[1]
  stock = row[2]
  key = oficina.lower() + str(sku)
  dict_stock[key] = stock
  dict_stock_no_asignado[key] = {'sku': sku, 'office': oficina, 'stock': stock}

# Por producir mes N
por_producir = load_workbook(filename_por_producir, data_only=True, read_only=True)
ws_por_producir = por_producir['Consolidado planificación']
max_row_por_producir = ws_por_producir.max_row
dict_producir = {}
dict_datos_modificados = {}

for row in ws_por_producir.iter_rows(min_row = 2, max_row = max_row_por_producir, values_only=True):
  key = row[0]                  # código SAP
  value = row[1]                # valor 
  dict_producir[key] = value
por_producir.close()

# Producción por despachar mes N 
por_despachar = load_workbook(filename_por_despachar, data_only=True, read_only=True)
ws_por_despachar = por_despachar['Din Conf-AP']
dict_por_despachar = {}
for row in ws_por_despachar.iter_rows(min_row = 5, max_row = ws_por_despachar.max_row, values_only=True):
  if row[0] is None:
    break
  llave = row[0].lower()
  kg_despachar = row[3]
  dict_por_despachar[llave] = kg_despachar
por_despachar.close()

# DELAY
for row in ws_delay.iter_rows(min_row = 15, max_row = ws_delay.max_row, values_only=True):
  if row[0] is None:
    break
  llave = row[0].lower()
  delay = row[3]
  dict_delay[llave] = delay
wb_stock.close()

# Vol. prom. Por contenedor
wb_vol_cont = load_workbook(filename_vol_cont, data_only=True, read_only=True)
ws_contenedor = wb_vol_cont['Volumen - Contenedor']
dict_vol_cont = {}

for row in ws_contenedor.iter_rows(min_row = 2, max_row = ws_contenedor.max_row, values_only=True):
  if row[0] is None:
    break
  llave = row[0].lower()
  volumen = row[3]
  dict_vol_cont[llave] = volumen
wb_vol_cont.close()

# En puerto a facturar
wb_puerto = load_workbook(filename_puerto, data_only=True, read_only=True)
ws_puerto = wb_puerto['Puerto']
dict_puerto = {}

for row in ws_puerto.iter_rows(min_row = 6, max_row = ws_puerto.max_row, values_only=True):
  if row[0] is None:
    break
  llave = row[0].lower()
  puerto_total = row[3]
  dict_puerto[llave] = puerto_total
wb_puerto.close()

# Agregamos todo a la planilla
for row in range(2, ws_irrestricto.max_row + 1):
  llave_actual = ws_irrestricto.cell(row = row, column = 1).value.lower()         # Llave = 'Agro Mexico1012764'
  sector = ws_irrestricto.cell(row = row, column = 2).value                       # Sector = 'Pollo'
  oficina = ws_irrestricto.cell(row = row, column = 3).value                      # oficina = 'Agro Mexico'
  sku_actual = ws_irrestricto.cell(row = row, column = 4).value                   # sku = '1012764'
  
  # % Util. prod
  concanate = (sector + oficina).lower()
  if concanate in util_dicc:
    ws_irrestricto[f'J{row}'].value = util_dicc[concanate]      # %Util.prod = '46%'
  elif concanate not in util_dicc:
    ws_irrestricto[f'J{row}'].value = parametro_porcentaje
    ws_irrestricto[f'J{row}'].fill = PatternFill("solid", fgColor=red)
    dict_datos_modificados[row] = {'llave': llave_actual, 'column': 10, 'name': '% Uti. producción', 'original_value': 0, 'change_value': 0.35 }

  # Stock
  if llave_actual in dict_stock:
    dict_stock_no_asignado.pop(llave_actual, None)
    ws_irrestricto[f'L{row}'].value = dict_stock[llave_actual]
  elif llave_actual not in dict_stock:
    ws_irrestricto[f'L{row}'].value = 0 
    ws_irrestricto[f'L{row}'].fill = PatternFill("solid", fgColor=red)
    dict_datos_modificados[row] = {'llave': llave_actual, 'column': 12, 'name': 'Stock al día', 'original_value': 'None', 'change_value': 0 }

# ------ Stock sin ventas ni producción ------
j = ws_irrestricto.max_row
for key, value in dict_stock_no_asignado.items():
  if value['office'].lower() in dict_tipo_venta[seleccion_tipo_venta]:
    if value['office'].lower() == 'agro mexico':
      nivel_2 = dict_materiales[str(value['sku'])]['nivel_2']
      if nivel_2 == 'carne recuperada':
        j += 1
        ws_irrestricto[f'A{j}'].value = key
        ws_irrestricto[f'C{j}'].value = value['office']
        ws_irrestricto[f'D{j}'].value = value['sku']
        ws_irrestricto[f'H{j}'].value = 0
        ws_irrestricto[f'I{j}'].value = 0
        ws_irrestricto[f'L{j}'].value = value['stock']
    else:
      j += 1
      ws_irrestricto[f'A{j}'].value = key
      ws_irrestricto[f'C{j}'].value = value['office']
      ws_irrestricto[f'D{j}'].value = value['sku']
      ws_irrestricto[f'H{j}'].value = 0
      ws_irrestricto[f'I{j}'].value = 0
      ws_irrestricto[f'L{j}'].value = value['stock']

    sector = ''

    if str(value['sku']) in dict_materiales.keys():
      sector = dict_materiales[str(value['sku'])]['sector']
      ws_irrestricto[f'B{j}'].value = sector
      ws_irrestricto[f'E{j}'].value = dict_materiales[str(value['sku'])]['descripcion']
      ws_irrestricto[f'F{j}'].value = dict_materiales[str(value['sku'])]['nivel_2']
      ws_irrestricto[f'G{j}'].value = dict_materiales[str(value['sku'])]['nivel_3']
    
    llave_util = sector.lower() + value['office'].lower()
    if llave_util in util_dicc:
      ws_irrestricto[f'J{j}'].value = util_dicc[llave_util]
    else:
      ws_irrestricto[f'J{j}'].value = parametro_porcentaje
      ws_irrestricto[f'J{j}'].fill = PatternFill("solid", fgColor=red)
      dict_datos_modificados[j] = {'llave': key, 'column': 10, 'name': '% Uti. producción', 'original_value': 0, 'change_value': 0.35 }

plan_irrestricto.save(filename)

# ------
for i in range(2, ws_irrestricto.max_row + 1):
  llave_actual = ws_irrestricto[f'A{i}'].value.lower() 
  sector = ws_irrestricto[f'B{i}'].value
  sku_actual = ws_irrestricto[f'D{i}'].value 

  # Por producir mes N
  if sku_actual in dict_producir:
    ponderacion = dict_ponderacion_cumplimiento[sector.lower()]
    ws_irrestricto[f'M{i}'].value = (float(dict_producir[sku_actual])* ponderacion)
    dict_producir.pop(sku_actual, None)
  elif sku_actual not in dict_producir:
    ws_irrestricto[f'M{i}'].value = 0
    ws_irrestricto[f'M{i}'].fill = PatternFill("solid", fgColor=red)
    dict_datos_modificados[i] = {'llave': llave_actual, 'column': 13, 'name': 'Por producir mes N', 'original_value': 'None', 'change_value': 0 }
  if llave_actual in dict_por_despachar:
    ws_irrestricto[f'N{i}'].value = dict_por_despachar[llave_actual]
  
  if llave_actual in dict_delay:
    ws_irrestricto[f'P{i}'].value = dict_delay[llave_actual]
  elif llave_actual not in dict_delay:
    ws_irrestricto[f'P{i}'].value = 0

  if llave_actual in dict_vol_cont:
    ws_irrestricto[f'Q{i}'].value = dict_vol_cont[llave_actual]
  elif llave_actual not in dict_vol_cont:
    ws_irrestricto[f'Q{i}'].value = 24000
    ws_irrestricto[f'Q{i}'].fill = PatternFill("solid", fgColor=red)
    dict_datos_modificados[i] = {'llave': llave_actual, 'column': 17, 'name': 'Vol. prom. Por contenedor', 'original_value': 'None', 'change_value': 24000 }

  # En puerto a facturar
  if llave_actual in dict_puerto:
    ws_irrestricto.cell(row = row, column = 29).value = dict_puerto[llave_actual]
plan_irrestricto.save(filename)

# ------ Por producir sin stock sin ventas ni producción ------
# for key, value in dict_producir.items():
#   if value > 0:
#     print(key, value)

# 4. Agregamos las columnas con formulas
for i, row in enumerate(ws_irrestricto.iter_rows(min_row = 2, max_row = ws_irrestricto.max_row, values_only=True), start = 2):
  # Prod. disponible
  prod = row[7] or 0                                            # Columna H
  util = row[9]                                                 # Columna J /% Uti. producción
  prod_disp = prod * util                                       # Columna K
  ws_irrestricto[f'K{i}'].value = prod_disp

  # Total disponible
  stock = row[11] or 0                                          # Columna L
  por_producir = row[12]                                        # Columna M
  produccion_por_despachar = row[13] or 0                       # Columna N
  total_disp = (prod_disp + stock + por_producir) - produccion_por_despachar
  ws_irrestricto[f'O{i}'].value = total_disp                            

  # Atraso a facturar
  delay = row[15]                                               # Columna P
  vol_contenedor = row[16] or 1                                 # Columna Q 
  if total_disp >= delay and delay > 0:
    # atrasos = int(delay / vol_contenedor)
    # atraso_factu = atrasos * vol_contenedor
    # ws_irrestricto.cell(row = i, column = 18).value = atraso_factu
    ws_irrestricto[f'R{i}'].value = delay
  
  elif total_disp < delay and delay > 0:
    # ws_irrestricto.cell(row = i, column = 18).value = delay
    atrasos = int(delay / vol_contenedor)
    atraso_factu = atrasos * vol_contenedor
    ws_irrestricto[f'R{i}'].value = atraso_factu
  
  else:
    ws_irrestricto[f'R{i}'].value = 0

plan_irrestricto.save(filename)

for i, row in enumerate(ws_irrestricto.iter_rows(min_row = 2, max_row = ws_irrestricto.max_row, values_only=True), start = 2):
  # Facturación atraso
  atraso_a_facturar = row[17] or 0                              # Columna R
  ajuste_atraso = row[18] or 0                                  # Columna S
  facturacion_atraso = atraso_a_facturar + ajuste_atraso
  ws_irrestricto[f'T{i}'].value = f'=R{i}-S{i}'
  
  # Producción para venta nueva
  total_disp = row[14]                                          # Columna O
  prd_venta_nueva = float(total_disp) - float(facturacion_atraso)
  ws_irrestricto[f'U{i}'].value = f'=O{i}-T{i}'

  # Venta del mes
  vol_contenedor = row[16] or 24000                             # Columna Q 
  venta_mes_n1 = row[8] or 0                                    # Columna J
  venta_mes = 0
  if prd_venta_nueva > vol_contenedor and venta_mes_n1 >= vol_contenedor:
    cant_contenedor = int(prd_venta_nueva / vol_contenedor)
    venta_mes = cant_contenedor * vol_contenedor                # Columna V
    ws_irrestricto.cell(row = i, column = 22).value = venta_mes

  else: ws_irrestricto.cell(row = i, column = 22).value = 0

  # Facturación Venta nueva                                      
  ajuste_venta_nueva = row[22] or 0                             # Columna W 
  facturacion_venta_nueva = 0
  if venta_mes > 0:
    ws_irrestricto.cell(row = i, column = 24).value = f'=V{i}+W{i}'
    facturacion_venta_nueva = venta_mes + ajuste_venta_nueva
  else: 
    ws_irrestricto.cell(row = i, column = 24).value = 0

  # Saldo Volumen disponible
  saldo_disp = prd_venta_nueva - facturacion_venta_nueva
  ws_irrestricto.cell(row = i, column = 25).value = f'=U{i}-X{i}'

  # Disponible stock sin venta
  delay = row[15]
  disponible_sin_venta = 0
  if saldo_disp > vol_contenedor and saldo_disp > venta_mes_n1 and delay < total_disp:
    cant_disp = int((saldo_disp + ajuste_venta_nueva) / vol_contenedor)
    disponible_sin_venta = cant_disp * vol_contenedor

    ws_irrestricto.cell(row = i, column = 26).value = disponible_sin_venta

  # Facturación stock
  ajuste_sin_venta = row[26] or 0                                   # Columna AA
  disponible_sin_venta = disponible_sin_venta or 0
  oficina = row[2]
  if disponible_sin_venta > 0:
    d_sin_venta = disponible_sin_venta + ajuste_sin_venta
    ws_irrestricto.cell(row = i, column = 28).value = f'=Z{i}+AA{i}'

  # Plan Irrestricto Inicial
  puerto_a_facturar = row[28] or 0
  if total_disp > 0:
    celda_plan_irrestricto = atraso_a_facturar + venta_mes + disponible_sin_venta + puerto_a_facturar
    ws_irrestricto.cell(row = i, column = 30).value = celda_plan_irrestricto

  # Plan ajustado
  if celda_plan_irrestricto > 0:
    # facturacion atraso + facturación venta nueva + facturación stock + En puerto a facturar
    ws_irrestricto.cell(row = i, column = 31).value = f'=T{i}+X{i}+AB{i}+AC{i}'

# Fila suma total
n_max = ws_irrestricto.max_row
ws_irrestricto.append({7: 'Total', 
  8: f'=SUM(H2:H{n_max})', 
  9: f'=SUM(I2:I{n_max})', 
  10: f'=AVERAGE(J2:J{n_max})',     # % Uti. producción
  11: f'=SUM(K2:K{n_max})', 
  12: f'=SUM(L2:L{n_max})', 
  13: f'=SUM(M2:M{n_max})',         # Por producir mes N
  14: f'=SUM(N2:N{n_max})',         # Producción por despachar mes N
  15: f'=SUM(O2:O{n_max})',         # Total disponible
  16: f'=SUM(P2:P{n_max})',         # Delay
  17: f'=AVERAGE(Q2:Q{n_max})',     # Vol. prom. Por contenedor
  18: f'=SUM(R2:R{n_max})',         # Atraso a facturar
  19: f'=SUM(S2:S{n_max})',         # Ajuste atraso
  20: f'=SUM(T2:T{n_max})',         # Facturación atraso
  21: f'=SUM(U2:U{n_max})',         # Producción para venta nueva
  22: f'=SUM(V2:V{n_max})',         # Venta del mes
  23: f'=SUM(W2:W{n_max})',         # Ajuste venta nueva
  24: f'=SUM(X2:X{n_max})',         # Facturación Venta nueva
  25: f'=SUM(Y2:Y{n_max})',         # Saldo Volumen disponible
  26: f'=SUM(Z2:Z{n_max})',         # Disponible stock sin venta
  27: f'=SUM(AA2:AA{n_max})',       # Ajuste stock sin venta
  28: f'=SUM(AB2:AB{n_max})',       # Facturación stock
  29: f'=SUM(AC2:AC{n_max})',       # En puerto a facturar
  30: f'=SUM(AD2:AD{n_max})',       # Plan Irrestricto
  31: f'=SUM(AE2:AE{n_max})',       # Plan Ajustado
  })

run_styles(ws_irrestricto, ws_irrestricto.max_row, ws_irrestricto.max_column)
date = datetime.now() + relativedelta(months = 1)
date_CL = month_translate_EN_CL[date.strftime('%B')]
plan_irrestricto.save(filename)
plan_irrestricto.close()
print("--- %s seconds ---" % (time.time() - start_time))
messageBox(dict_datos_modificados, dict_otro_tipo, path_img)
