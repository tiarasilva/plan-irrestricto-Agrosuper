from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE, BUILTIN_FORMATS
from constants import *

# plan_irrestricto = load_workbook(filename)
# ws = plan_irrestricto.active
# max_row = ws.max_row

## 2.1 Para los estilos
def run_styles(ws, row_max, col_max):
  thin = Side(border_style="thin", color=white)
  for col in range(1, col_max + 1):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color=white)  # tomamos la primera fila
    ws[get_column_letter(col) + '1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws[get_column_letter(col) + '1'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    nombre = ws.cell(row = 1, column = col).value
    ws.column_dimensions[get_column_letter(col)].width = tamano[nombre]
    ws.row_dimensions[1].height = 30
    ws[f'{get_column_letter(col)}{row_max}'].border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws[f'{get_column_letter(col)}{row_max}'].number_format = BUILTIN_FORMATS[3]
    ws[f'{get_column_letter(col)}{row_max}'].fill = PatternFill("solid", fgColor=blue)
    ws[f'{get_column_letter(col)}{row_max}'].font = Font(bold=True, color=white)

  for col in range(1, 18):
    ws[get_column_letter(col) + '1'].fill = PatternFill("solid", fgColor=grey)

  for col in range(18, 29, 4):
    ws[get_column_letter(col) + '1'].fill = PatternFill("solid", fgColor=lightBlue)
    ws[get_column_letter(col + 1) + '1'].fill = PatternFill("solid", fgColor=lightBlue)
    ws[get_column_letter(col + 2) + '1'].fill = PatternFill("solid", fgColor=orange)
    ws[get_column_letter(col + 3) + '1'].fill = PatternFill("solid", fgColor=grey)

  ws[get_column_letter(29) + '1'].fill = PatternFill("solid", fgColor=orange)
  ws[get_column_letter(30) + '1'].fill = PatternFill("solid", fgColor=blue)
  ws[get_column_letter(31) + '1'].fill = PatternFill("solid", fgColor=blue)
  ws[get_column_letter(32) + '1'].fill = PatternFill("solid", fgColor=blue)

  for colu in range(8, 32):
    letter = str(get_column_letter(colu))
    if letter != 'J' or letter != 'AF':
      col = ws.column_dimensions[letter]
      col.number_format = BUILTIN_FORMATS[3]

  for i in range(2, row_max + 1):
    # Colores rojo ajustes
    ws[f'S{i}'].font = Font(bold=True, color=red)
    ws[f'W{i}'].font = Font(bold=True, color=red)
    ws[f'AA{i}'].font = Font(bold=True, color=red)

    # FORMATOS
    ws[f'H{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'I{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'J{i}'].number_format = FORMAT_PERCENTAGE
    ws[f'K{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'L{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'M{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'N{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'O{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'P{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'Q{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'R{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'S{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'T{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'U{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'V{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'W{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'X{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'Y{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'Z{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'AA{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'AB{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'AC{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'AD{i}'].number_format = BUILTIN_FORMATS[3]
    ws[f'AE{i}'].number_format = BUILTIN_FORMATS[3]


# plan_irrestricto.close()